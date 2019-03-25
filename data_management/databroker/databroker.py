# encoding: utf-8
import importlib
import argparse
import xlrd,xlwt
from openpyxl import Workbook
from xlrd import xldate_as_tuple # 处理excel单元格的date数据
import time
import requests
import json
import collections
from datetime import datetime
from tqdm import tqdm

from log_module import log


class databroker(object):
    def __init__(self,*args,**kwargs):
        self.url = kwargs["url"]
        self.corpus = kwargs["corpus"]
        self.processor = []
        for arg in args:
            mod = importlib.import_module('databroker.databroker_module.{}'.format(arg))
            process_obj = getattr(mod,arg)()
            self.processor.append(process_obj)

    def Standardize(self,desc_id,output=None,Sequence=False,position=1):
        """
        param: desc_id: task_id and desc_id
        return result
        result:
        [
            {"query":...,tags:[
                               {"key":...,"value":...,"desc_id":...,"time":...},
                               {"key":...,"value":...,"desc_id":...,"time":...},
                               ...
                               ]},
            {"query":...,tags:[
                               {"key":...,"value":...,"desc_id":...,"time":...},
                               {"key":...,"value":...,"desc_id":...,"time":...},
                               ...
                               ]},
        ]
        """
        # log writter
        log_writter = log.log().getLog()

        if desc_id == None:
            desc_id = time.strftime('%Y-%m-%d-%H-%M-%S',time.localtime(time.time()))
        result = []
        data = xlrd.open_workbook(self.corpus)
        table = data.sheet_by_index(0)
        headers = [ table.cell(0,i).value for i in range(table.ncols) ]
        state = "{}"
        for i in tqdm(range(1,table.nrows),position=position,desc="Thread #{}".format(str(position))):
            query = str(table.cell(i,0).value)
            element = collections.OrderedDict()
            element["query"] = query
            element["tags"] = self.__getOrigintagsbyrow(table,i,headers,desc_id)
            if len(self.processor) == 0:
                result.append(element)
                log_writter.info(element["tags"])
            else:
                if len(query) == 0:
                    element["tags"].extend(self.__getData("query_empty",desc_id))
                    result.append(element)
                    log_writter.info("{} fail: {} ".format(query,"query_empty"))
                    continue
                post_data = {
                'query' : query,
                'state' : state
                }
                try:
                    r = requests.post(self.url, json=post_data,timeout=10)
                except Exception as e:
                    element["tags"].extend(self.__getData("timeout",desc_id))
                    result.append(element)
                    log_writter.info("{} fail: {} ".format(query,"timeout"))
                    state = "{}"
                    continue
                if("msg" in r.json()):
                    msg = r.json()['msg']
                    if Sequence:
                        state = msg["state"]
                    element["tags"].extend(self.__getData("process",desc_id,msg=msg))
                    result.append(element)
                    log_writter.info("{} success ".format(query))
                else:
                    element["tags"].extend(self.__getData("response_illegal",desc_id))
                    result.append(element)
                    log_writter.info("{} fail: {} ".format(query,"response_illegal"))
                    state = "{}"
        # 为每条语料打上status标记，有标注的标记为1,没有标注的记的为0
        for res in result:
            tag = collections.OrderedDict()
            tag["key"] = "status"
            tag["value"] = 0
            if len(res["tags"]) != 0:
                tag["value"] = 1
            tag["desc_id"] = desc_id
            tag["time"] = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
            res["tags"].append(tag)
        if output is not None:
            with open(output,"w",encoding="utf-8") as fd:
                json.dump(result,fd,ensure_ascii=False)
        standard_form = json.dumps(result,ensure_ascii=False)
        return standard_form
                
    def Specialize(self,standard_form,filename):
        data_json = json.loads(standard_form,object_pairs_hook=collections.OrderedDict,encoding="utf-8")
        output = Workbook()
        output_sheet = output.active
        headers = ['query']
        headers_tags = [ele["key"] for ele in data_json[0]["tags"]]
        headers.extend(headers_tags)
        output_sheet.append(headers)
        for data in data_json:
            tmp = []
            tmp.append(data["query"])
            for tag in data["tags"]:
                if tag["value"] in ['#NULL!','#DIV/0!','#VALUE!','#REF!', '#NAME?','#NUM!', '#N/A']:
                    tmp.append(tag["value"].lstrip('#'))
                else:
                    tmp.append(tag["value"])
            output_sheet.append(tmp)
        return output.save(filename)

        
    def __getOrigintagsbyrow(self,table,row,headers,desc_id):
        origin_tags = []
        for j in range(1,table.ncols):
            element = collections.OrderedDict()
            element["key"] = headers[j]
            # cell_type 对 XL_CELL_ERROR 和 XL_CELL_DATE 的处理
            if table.cell_type(row,j) == 5: # XL_CELL_ERROR
                element["value"] = str(xlrd.biffh.error_text_from_code[table.cell(row,j).value])
            elif table.cell_type(row,j) == 3: # XL_CELL_DATE
                date_tuple = xldate_as_tuple(table.cell(row,j).value,0)
                date_list = list(date_tuple)
                date_tuple = self.__processInvalidDatetuple(date_list)
                element["value"] = str(datetime(*date_list))
            else:
                element["value"] = str(table.cell(row,j).value)
            element["desc_id"] = desc_id            
            element["time"] = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
            origin_tags.append(element)
        return origin_tags
        
    
    def __getData(self,label,desc_id,msg=None):
        """
        label: the string of the name of databroker_module's method 
        """
        result = [] 
        for process in self.processor:
            tag = collections.OrderedDict()
            method = getattr(process,label)
            if label == "process":
                tag["key"],tag["value"] = method(msg)
            else:
                tag["key"],tag["value"] = method()
            tag["desc_id"] = desc_id
            tag["time"] = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
            result.append(tag)
        return result
    
    def __processInvalidDatetuple(self,date_list):
        '''
        the 1st,2nd,3rd data in date_list must be greater than 0,the rest does not meet the requirements
        '''
        for i in range(3):
            if date_list[i] == 0:
                date_list[i] = 1
        return date_list
    
        

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='databroker')
    parser.add_argument('--host',default='115.182.62.171')
    parser.add_argument('-P','--port',default='1781')
    parser.add_argument('-b','--bot')
    parser.add_argument('-c','--corpus')
    args = parser.parse_args()
    host = args.host
    port = args.port
    bot = args.bot
    url  = 'http://{host}:{port}/bot/{bot}/simulator'.format(host=host,port=port,bot=bot)
    corpus = args.corpus
    d = databroker("entity","intent",url=url,corpus=corpus)
    result_json = d.Standardize("task01_batch01")
    #print(result_json)
    d.Specialize(result_json,"output1.xlsx")
    
